"""Stream large .xlsx files into SQLite without loading the whole workbook as a DataFrame."""

from __future__ import annotations

import json
import math
import os
import re
import sqlite3
from datetime import date, datetime
from numbers import Number
from typing import Callable, Optional, Sequence

from dates import normalize_cell_date, normalize_cell_text
from schema import (
    COLUMNS,
    KPI_FIELDS,
    KPI_SQL,
    LABEL_TO_SQL,
    NUMERIC_FIELDS,
    SEARCH_PRIORITY,
    SQL_COLUMNS,
    sql_name,
)

ProgressFn = Optional[Callable[[int, str], None]]

HEADER_SCAN_ROWS = 15
MIN_HEADER_HITS = 8
INSERT_BATCH = 800
MAX_ROWS = 2_000_000


def normalize_header(value) -> str:
    text = str(value if value is not None else "")
    text = text.replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip().lower()


def empty_kpis() -> dict:
    kpis = {key: 0.0 for key in KPI_FIELDS}
    kpis["invoiceCount"] = 0
    return kpis


def cache_path(cache_dir: str, fingerprint: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", fingerprint)[:120]
    return os.path.join(cache_dir, safe + ".sqlite")


def open_db(db_path: str) -> sqlite3.Connection:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    con.create_function("dsr_date", 1, normalize_cell_date, deterministic=True)
    con.create_function("dsr_text", 1, normalize_cell_text, deterministic=True)
    return con


def is_ready(db_path: str) -> bool:
    if not os.path.isfile(db_path):
        return False
    try:
        con = sqlite3.connect(db_path)
        try:
            row = con.execute(
                "SELECT value FROM meta WHERE key = 'ready'"
            ).fetchone()
            return bool(row and row[0] == "1")
        finally:
            con.close()
    except sqlite3.Error:
        return False


def ingest_file(src_path: str, db_path: str, progress: ProgressFn = None) -> dict:
    """Parse src_path into db_path. Returns KPI payload."""
    _emit(progress, 4, "Opening workbook…")
    os.makedirs(os.path.dirname(db_path) or ".", exist_ok=True)
    tmp_path = db_path + ".tmp"
    for path in (tmp_path, db_path):
        if os.path.exists(path):
            os.remove(path)

    con = sqlite3.connect(tmp_path)
    con.execute("PRAGMA journal_mode = WAL")
    con.execute("PRAGMA synchronous = OFF")
    con.execute("PRAGMA temp_store = MEMORY")
    try:
        try:
            stats, sheet_name, spec = _ingest_calamine(src_path, con, progress)
        except Exception:
            _emit(progress, 8, "Using streaming Excel reader…")
            stats, sheet_name, spec = _ingest_openpyxl(src_path, con, progress)
        _write_meta(con, src_path, sheet_name, stats, spec)
        stats["columns"] = spec["labels"]
        con.commit()
    except Exception:
        con.close()
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    con.close()
    os.replace(tmp_path, db_path)
    _emit(progress, 100, f"Ready · {stats['invoiceCount']:,} invoices")
    return stats


def query_rows(
    db_path: str,
    start: int,
    end: int,
    search: str = "",
    sort_col: str = "",
    sort_dir: str = "asc",
    filters: dict | None = None,
    grid_filters: dict | None = None,
) -> dict:
    start = max(0, int(start))
    end = max(start, int(end))
    limit = min(500, max(1, end - start))
    con = open_db(db_path)
    try:
        meta = {r[0]: r[1] for r in con.execute("SELECT key, value FROM meta")}
        labels, sql_cols, label_to_sql = _schema_from_meta(meta)
        where, params = compose_where(search, sql_cols, label_to_sql, filters)
        where_g, params_g = _grid_filter_where(label_to_sql, grid_filters)
        where, params = _merge_where(where, params, where_g, params_g)
        order = _order_clause(sort_col, sort_dir, label_to_sql, sql_cols)
        total = con.execute(
            f"SELECT COUNT(*) FROM invoices {where}", params
        ).fetchone()[0]
        col_sql = ", ".join(f'"{c}"' for c in sql_cols)
        rows_raw = con.execute(
            f"SELECT {col_sql} FROM invoices {where} {order} LIMIT ? OFFSET ?",
            [*params, limit, start],
        ).fetchall()
        rows = []
        for raw in rows_raw:
            rec = {}
            for label, sql in zip(labels, sql_cols):
                rec[label] = _json_cell(raw[sql])
            rows.append(rec)
        return {
            "rows": rows,
            "columns": labels,
            "lastRow": int(total),
            "filtered": bool(where),
        }
    finally:
        con.close()


def read_summary(db_path: str) -> dict:
    con = sqlite3.connect(db_path)
    try:
        meta = {r[0]: r[1] for r in con.execute("SELECT key, value FROM meta")}
    finally:
        con.close()
    labels, _, _ = _schema_from_meta(meta)
    return {
        "kpis": _kpis_from_meta(meta),
        "rowCount": int(meta.get("row_count") or 0),
        "sheetName": meta.get("sheet_name") or "",
        "sourceName": meta.get("source_name") or "",
        "columns": labels,
    }


def _ingest_calamine(path: str, con: sqlite3.Connection, progress: ProgressFn):
    from python_calamine import CalamineWorkbook

    _emit(progress, 8, "Reading Excel with high-speed parser…")
    wb = CalamineWorkbook.from_path(path)
    for name in wb.sheet_names:
        sheet = wb.get_sheet_by_name(name)
        values = sheet.to_python(skip_empty_area=True)
        located = _locate_header(values)
        if not located:
            continue
        located["rows"] = values
        spec = _build_spec(located)
        _create_tables(con, spec)
        stats = _write_rows(con, spec, progress)
        return stats, name, spec
    raise ValueError("Could not find the Daily Sales Register column headers in this file.")


def _ingest_openpyxl(path: str, con: sqlite3.Connection, progress: ProgressFn):
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    try:
        for sheet in wb.worksheets:
            located = None
            header_row = None
            spec = None
            stats = empty_kpis()
            batch = []
            written = 0
            insert_sql = None

            for i, raw in enumerate(sheet.iter_rows(values_only=True)):
                line = list(raw)
                if located is None:
                    if i >= HEADER_SCAN_ROWS:
                        break
                    probe = _locate_header([line])
                    if probe:
                        located = probe
                        header_row = i
                        spec = _build_spec({"rows": [line], "header_row": 0, "index_map": probe["index_map"]})
                        _create_tables(con, spec)
                        placeholders = ",".join(["?"] * len(spec["sql_cols"]))
                        col_list = ",".join(f'"{c}"' for c in spec["sql_cols"])
                        insert_sql = f"INSERT INTO invoices ({col_list}) VALUES ({placeholders})"
                    continue
                if i <= header_row:
                    continue
                if _is_empty(line):
                    continue
                rec = _map_row(line, spec)
                if rec is None:
                    continue
                batch.append(rec)
                _add_kpis(stats, rec, spec["labels"])
                written += 1
                if len(batch) >= INSERT_BATCH:
                    con.executemany(insert_sql, batch)
                    batch.clear()
                    _emit(progress, min(95, 12 + written // 4000), f"Indexing invoices… {written:,}")
                if written >= MAX_ROWS:
                    break

            if located and batch:
                con.executemany(insert_sql, batch)
            if located and written:
                stats["invoiceCount"] = written
                for key in KPI_FIELDS:
                    stats[key] = round(float(stats[key]), 2)
                return stats, sheet.title, spec
        raise ValueError("Could not find the Daily Sales Register column headers in this file.")
    finally:
        wb.close()


def _locate_header(rows: Sequence[Sequence]):
    expected = {normalize_header(c): i for i, c in enumerate(COLUMNS)}
    best = None
    best_score = 0
    scan = min(len(rows), HEADER_SCAN_ROWS)
    for r in range(scan):
        line = rows[r] if r < len(rows) else []
        index_map = {}
        for col_idx, cell in enumerate(line):
            key = normalize_header(cell)
            if key in expected and COLUMNS[expected[key]] not in index_map:
                index_map[COLUMNS[expected[key]]] = col_idx
        score = len(index_map)
        if score > best_score:
            best_score = score
            best = {"header_row": r, "index_map": index_map, "rows": rows}
    if not best or best_score < MIN_HEADER_HITS:
        return None
    return best


def _build_spec(located: dict) -> dict:
    rows = located["rows"]
    header_row_idx = located["header_row"]
    header_line = rows[header_row_idx] if header_row_idx < len(rows) else []
    index_map = dict(located["index_map"])
    labels = list(COLUMNS)
    used = set(index_map.values())
    for i, cell in enumerate(header_line):
        if i in used or cell is None or cell == "":
            continue
        display = str(cell).replace("\xa0", " ")
        display = re.sub(r"[\r\n]+", " ", display)
        display = re.sub(r"\s+", " ", display).strip()[:120]
        if not display or display in index_map:
            continue
        labels.append(display)
        index_map[display] = i
    sql_cols = []
    seen = set()
    for label in labels:
        base = sql_name(label)
        sql = base
        n = 2
        while sql in seen:
            sql = f"{base}_{n}"
            n += 1
        seen.add(sql)
        sql_cols.append(sql)
    return {
        "labels": labels,
        "sql_cols": sql_cols,
        "label_to_sql": dict(zip(labels, sql_cols)),
        "index_map": index_map,
        "header_row": header_row_idx,
        "rows": rows,
    }


def _schema_from_meta(meta: dict):
    try:
        if meta.get("columns_json") and meta.get("sql_columns_json"):
            labels = json.loads(meta["columns_json"])
            sql_cols = json.loads(meta["sql_columns_json"])
            if labels and sql_cols and len(labels) == len(sql_cols):
                return labels, sql_cols, dict(zip(labels, sql_cols))
    except (json.JSONDecodeError, TypeError):
        pass
    return COLUMNS, SQL_COLUMNS, LABEL_TO_SQL


def _create_tables(con: sqlite3.Connection, spec: dict) -> None:
    cols = []
    for label, sql in zip(spec["labels"], spec["sql_cols"]):
        kind = "REAL" if label in NUMERIC_FIELDS else "TEXT"
        cols.append(f'"{sql}" {kind}')
    con.execute("DROP TABLE IF EXISTS invoices")
    con.execute("DROP TABLE IF EXISTS meta")
    con.execute(f"CREATE TABLE invoices ({', '.join(cols)})")
    con.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)")
    l2s = spec["label_to_sql"]
    if "Invoice Number" in l2s:
        con.execute(f'CREATE INDEX idx_invoice ON invoices ("{l2s["Invoice Number"]}")')
    if "Advertiser Name" in l2s:
        con.execute(f'CREATE INDEX idx_advertiser ON invoices ("{l2s["Advertiser Name"]}")')


def _write_rows(con: sqlite3.Connection, spec: dict, progress: ProgressFn) -> dict:
    rows = spec["rows"]
    header_row = spec["header_row"]
    placeholders = ",".join(["?"] * len(spec["sql_cols"]))
    col_list = ",".join(f'"{c}"' for c in spec["sql_cols"])
    insert_sql = f"INSERT INTO invoices ({col_list}) VALUES ({placeholders})"

    kpis = empty_kpis()
    batch = []
    data_rows = rows[header_row + 1 :]
    total_guess = max(len(data_rows), 1)
    written = 0

    for i, raw in enumerate(data_rows):
        if written >= MAX_ROWS:
            break
        if _is_empty(raw):
            continue
        rec = _map_row(raw, spec)
        if rec is None:
            continue
        batch.append(rec)
        _add_kpis(kpis, rec, spec["labels"])
        written += 1
        if len(batch) >= INSERT_BATCH:
            con.executemany(insert_sql, batch)
            batch.clear()
            pct = 12 + int(82 * min(i / total_guess, 1))
            _emit(progress, pct, f"Indexing invoices… {written:,}")

    if batch:
        con.executemany(insert_sql, batch)
    kpis["invoiceCount"] = written
    for key in KPI_FIELDS:
        kpis[key] = round(float(kpis[key]), 2)
    return kpis


def _map_row(raw: Sequence, spec: dict) -> Optional[tuple]:
    values = []
    any_value = False
    index_map = spec["index_map"]
    for label in spec["labels"]:
        idx = index_map.get(label)
        cell = raw[idx] if idx is not None and idx < len(raw) else None
        parsed = _coerce(cell, label)
        if parsed not in ("", None):
            any_value = True
        values.append(parsed)
    return tuple(values) if any_value else None


def _coerce(value, label: str):
    if value is None or value == "":
        return None if label in NUMERIC_FIELDS else ""
    if label in ("Invoice Date", "Service Period", "Bank Realization date", "Billing Month", "Collection month"):
        iso = normalize_cell_date(value)
        if iso:
            return iso
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if label in NUMERIC_FIELDS:
        return _to_number(value)
    if isinstance(value, Number) and not isinstance(value, bool):
        num = float(value)
        if math.isfinite(num) and num == int(num) and abs(num) < 1e15:
            return str(int(num))
        return str(value).strip()
    text = str(value).replace("\x00", "").strip()
    return text[:8000]


def _to_number(value) -> Optional[float]:
    if isinstance(value, Number) and not isinstance(value, bool):
        num = float(value)
        return num if math.isfinite(num) else None
    text = str(value).strip()
    if not text:
        return None
    neg = text.startswith("(") and text.endswith(")")
    text = re.sub(r"[₹$€£,\s()]", "", text)
    if not text or text == "-":
        return None
    try:
        num = float(text)
    except ValueError:
        return None
    if not math.isfinite(num):
        return None
    return -num if neg else num


def _add_kpis(kpis: dict, rec: tuple, labels: Sequence[str]) -> None:
    for i, label in enumerate(labels):
        if label in KPI_FIELDS:
            val = rec[i]
            if isinstance(val, (int, float)) and math.isfinite(val):
                kpis[label] += float(val)


def _is_empty(raw: Sequence) -> bool:
    for cell in raw:
        if cell is None or cell == "":
            continue
        if isinstance(cell, str) and not cell.strip():
            continue
        return False
    return True


def _write_meta(con: sqlite3.Connection, src_path: str, sheet_name: str, stats: dict, spec: dict) -> None:
    pairs = {
        "ready": "1",
        "row_count": str(stats["invoiceCount"]),
        "sheet_name": sheet_name,
        "source_name": os.path.basename(src_path),
        "source_path": src_path,
        "columns_json": json.dumps(spec["labels"], ensure_ascii=False),
        "sql_columns_json": json.dumps(spec["sql_cols"]),
    }
    for key in KPI_FIELDS:
        pairs[f"kpi_{sql_name(key)}"] = str(stats[key])
    con.executemany(
        "INSERT OR REPLACE INTO meta(key, value) VALUES (?, ?)",
        list(pairs.items()),
    )


def _kpis_from_meta(meta: dict) -> dict:
    kpis = empty_kpis()
    kpis["invoiceCount"] = int(float(meta.get("row_count") or 0))
    for key in KPI_FIELDS:
        kpis[key] = float(meta.get(f"kpi_{sql_name(key)}") or 0)
    return kpis


def _kpis_from_sql(con: sqlite3.Connection, where: str, params: list, total: int) -> dict:
    sums = ", ".join(f'COALESCE(SUM("{c}"), 0)' for c in KPI_SQL)
    row = con.execute(f"SELECT {sums} FROM invoices {where}", params).fetchone()
    kpis = empty_kpis()
    kpis["invoiceCount"] = total
    for i, key in enumerate(KPI_FIELDS):
        kpis[key] = round(float(row[i] or 0), 2)
    return kpis


def compose_where(search: str, sql_cols: Sequence[str], label_to_sql: dict, filters: dict | None = None):
    where_s, params_s = _where_clause(search, sql_cols)
    where_f, params_f = _filter_where(label_to_sql, filters or {})
    return _merge_where(where_s, params_s, where_f, params_f)


def _ident(label_to_sql: dict, label: str) -> str | None:
    sql = label_to_sql.get(label)
    if not sql:
        return None
    return '"' + str(sql).replace('"', "") + '"'


def _filter_where(label_to_sql: dict, filters: dict):
    clauses = []
    params = []
    date_field = filters.get("dateField") or "Invoice Date"
    if date_field not in ("Invoice Date", "Service Period"):
        date_field = "Invoice Date"
    date_col = _ident(label_to_sql, date_field)
    date_from = str(filters.get("dateFrom") or "").strip()
    date_to = str(filters.get("dateTo") or "").strip()
    if date_col and (re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_from) or re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_to)):
        expr = f"dsr_date({date_col})"
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_from) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_to):
            clauses.append(f"{expr} != '' AND {expr} BETWEEN ? AND ?")
            params.extend([date_from, date_to])
        elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_from):
            clauses.append(f"{expr} != '' AND {expr} >= ?")
            params.append(date_from)
        else:
            clauses.append(f"{expr} != '' AND {expr} <= ?")
            params.append(date_to)
    currency = str(filters.get("currency") or "").strip()
    cur_col = _ident(label_to_sql, "Currency")
    cur_key = normalize_cell_text(currency)
    if cur_key and cur_key != "all" and cur_col:
        clauses.append(f"dsr_text({cur_col}) = ?")
        params.append(cur_key[:40])
    brands = [normalize_cell_text(x)[:120] for x in (filters.get("brands") or []) if normalize_cell_text(x)][:40]
    brand_col = _ident(label_to_sql, "Brand Name")
    if brands and brand_col:
        clauses.append(f"dsr_text({brand_col}) IN ({','.join(['?'] * len(brands))})")
        params.extend(brands)
    teams = [normalize_cell_text(x)[:120] for x in (filters.get("teams") or []) if normalize_cell_text(x)][:40]
    team_col = _ident(label_to_sql, "Team Responsible")
    if teams and team_col:
        clauses.append(f"dsr_text({team_col}) IN ({','.join(['?'] * len(teams))})")
        params.extend(teams)
    if not clauses:
        return "", []
    return "WHERE " + " AND ".join(clauses), params


_DATE_FILTER_LABELS = {
    "Invoice Date",
    "Service Period",
    "Bank Realization date",
    "Billing Month",
    "Collection month",
}
_NUM_FILTER_LABELS = set(NUMERIC_FIELDS)


def _grid_filter_where(label_to_sql: dict, model) -> tuple:
    if not isinstance(model, dict) or not model:
        return "", []
    clauses = []
    params = []
    for i, (label, spec) in enumerate(model.items()):
        if i >= 40:
            break
        col = _ident(label_to_sql, str(label))
        if not col or not isinstance(spec, dict) or not spec:
            continue
        piece, piece_params = _ag_filter_clause(col, str(label), spec)
        if piece:
            clauses.append(piece)
            params.extend(piece_params)
    if not clauses:
        return "", []
    return "WHERE " + " AND ".join(clauses), params


def _ag_filter_clause(col: str, label: str, spec: dict) -> tuple:
    nested = spec.get("conditions")
    if not isinstance(nested, list):
        nested = [spec.get("condition1"), spec.get("condition2")]
        nested = [c for c in nested if isinstance(c, dict)]
    else:
        nested = [c for c in nested if isinstance(c, dict)]
    if nested:
        op = str(spec.get("operator") or "AND").upper()
        if op not in ("AND", "OR"):
            op = "AND"
        parts, params = [], []
        for cond in nested[:4]:
            piece, piece_params = _ag_simple_filter(col, label, cond)
            if piece:
                parts.append(piece)
                params.extend(piece_params)
        if not parts:
            return "", []
        if len(parts) == 1:
            return parts[0], params
        return "(" + f" {op} ".join(parts) + ")", params
    return _ag_simple_filter(col, label, spec)


def _like_fragment(value: str) -> str:
    text = str(value).replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return text[:200]


def _iso_day(value) -> str:
    text = str(value or "").strip()
    if len(text) >= 10 and re.fullmatch(r"\d{4}-\d{2}-\d{2}", text[:10]):
        return text[:10]
    return ""


def _ag_simple_filter(col: str, label: str, spec: dict) -> tuple:
    kind = str(spec.get("filterType") or "").lower()
    op = str(spec.get("type") or "").strip() or "contains"
    if kind == "set":
        values = [normalize_cell_text(v)[:120] for v in (spec.get("values") or []) if v is not None and str(v).strip()][:80]
        if not values:
            # Empty set filter is "not finished yet", not "match nothing".
            return "", []
        placeholders = ",".join(["?"] * len(values))
        return f"dsr_text({col}) IN ({placeholders})", values

    if op in ("blank", "empty"):
        return f"({col} IS NULL OR TRIM(CAST({col} AS TEXT)) IN ('', 'None'))", []
    if op in ("notBlank", "notEmpty"):
        return f"({col} IS NOT NULL AND TRIM(CAST({col} AS TEXT)) NOT IN ('', 'None'))", []

    if kind == "number" or (kind != "date" and label in _NUM_FILTER_LABELS):
        expr = f"CAST({col} AS REAL)"
        try:
            left = spec.get("filter")
            right = spec.get("filterTo")
            n1 = None if left in (None, "") else float(left)
            n2 = None if right in (None, "") else float(right)
        except (TypeError, ValueError):
            return "", []
        if op == "equals" and n1 is not None:
            return f"{expr} = ?", [n1]
        if op == "notEqual" and n1 is not None:
            return f"{expr} != ?", [n1]
        if op == "lessThan" and n1 is not None:
            return f"{expr} < ?", [n1]
        if op == "lessThanOrEqual" and n1 is not None:
            return f"{expr} <= ?", [n1]
        if op == "greaterThan" and n1 is not None:
            return f"{expr} > ?", [n1]
        if op == "greaterThanOrEqual" and n1 is not None:
            return f"{expr} >= ?", [n1]
        if op == "inRange" and n1 is not None and n2 is not None:
            lo, hi = (n1, n2) if n1 <= n2 else (n2, n1)
            return f"{expr} BETWEEN ? AND ?", [lo, hi]
        return "", []

    if kind == "date" or label in _DATE_FILTER_LABELS:
        expr = f"dsr_date({col})"
        d1 = _iso_day(spec.get("dateFrom") if spec.get("dateFrom") not in (None, "") else spec.get("filter"))
        d2 = _iso_day(spec.get("dateTo") if spec.get("dateTo") not in (None, "") else spec.get("filterTo"))
        if op == "equals" and d1:
            return f"{expr} = ?", [d1]
        if op == "notEqual" and d1:
            return f"{expr} != '' AND {expr} != ?", [d1]
        if op == "lessThan" and d1:
            return f"{expr} != '' AND {expr} < ?", [d1]
        if op == "lessThanOrEqual" and d1:
            return f"{expr} != '' AND {expr} <= ?", [d1]
        if op == "greaterThan" and d1:
            return f"{expr} != '' AND {expr} > ?", [d1]
        if op == "greaterThanOrEqual" and d1:
            return f"{expr} != '' AND {expr} >= ?", [d1]
        if op == "inRange" and d1 and d2:
            lo, hi = (d1, d2) if d1 <= d2 else (d2, d1)
            return f"{expr} != '' AND {expr} BETWEEN ? AND ?", [lo, hi]
        return "", []

    raw = spec.get("filter")
    if raw is None or isinstance(raw, (dict, list)) or str(raw).strip() == "":
        return "", []
    needle = normalize_cell_text(raw)[:200]
    if not needle:
        return "", []
    like = _like_fragment(needle)
    if op == "equals":
        return f"dsr_text({col}) = ?", [needle]
    if op == "notEqual":
        return f"dsr_text({col}) != ?", [needle]
    if op == "contains":
        return f"dsr_text({col}) LIKE ? ESCAPE '\\'", [f"%{like}%"]
    if op == "notContains":
        return f"dsr_text({col}) NOT LIKE ? ESCAPE '\\'", [f"%{like}%"]
    if op == "startsWith":
        return f"dsr_text({col}) LIKE ? ESCAPE '\\'", [f"{like}%"]
    if op == "endsWith":
        return f"dsr_text({col}) LIKE ? ESCAPE '\\'", [f"%{like}"]
    return f"dsr_text({col}) LIKE ? ESCAPE '\\'", [f"%{like}%"]


def _merge_where(where_a: str, params_a: list, where_b: str, params_b: list):
    parts = []
    params = []
    for where, par in ((where_a, params_a), (where_b, params_b)):
        text = (where or "").strip()
        if not text:
            continue
        if text.upper().startswith("WHERE "):
            text = text[6:]
        parts.append(f"({text})")
        params.extend(par)
    if not parts:
        return "", []
    return "WHERE " + " AND ".join(parts), params


def _where_clause(search: str, sql_cols: Sequence[str]) -> tuple:
    q = (search or "").strip()
    if not q:
        return "", []
    if len(q) > 200:
        q = q[:200]
    folded = normalize_cell_text(q)[:200]
    if not folded:
        return "", []
    like = f"%{_like_fragment(folded)}%"
    prefer = []
    seen = set()
    for label in SEARCH_PRIORITY:
        sql = LABEL_TO_SQL.get(label)
        if sql and sql in sql_cols and sql not in seen:
            prefer.append(sql)
            seen.add(sql)
    ordered = prefer + [c for c in sql_cols if c not in seen]
    ors = []
    params = []
    for col in ordered:
        ors.append(f'dsr_text("{col}") LIKE ? ESCAPE \'\\\'')
        params.append(like)
    return "WHERE " + " OR ".join(ors), params


def _order_clause(sort_col: str, sort_dir: str, label_to_sql: dict, sql_cols: Sequence[str]) -> str:
    sql = label_to_sql.get(sort_col) or (sort_col if sort_col in sql_cols else "")
    if not sql:
        return ""
    direction = "DESC" if str(sort_dir).lower() == "desc" else "ASC"
    return f'ORDER BY "{sql}" IS NULL, "{sql}" {direction}'


def _json_cell(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        if value == int(value) and abs(value) < 1e12:
            return int(value)
        return round(value, 4)
    return value


def _emit(progress: ProgressFn, pct: int, message: str) -> None:
    if progress:
        progress(max(0, min(100, pct)), message)
